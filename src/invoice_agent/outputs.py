"""File-system side effects for the agent: Excel append + review-folder dump.

Lives here (rather than inside `tools.py`) so the I/O is unit-testable on its
own and the tools file stays focused on the SDK glue.
"""

from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from invoice_agent.schemas import Invoice, ReviewReason

PROJECT_ROOT = Path(__file__).resolve().parents[2]
OUTPUT_DIR = PROJECT_ROOT / "output"
PROCESSED_XLSX = OUTPUT_DIR / "processed.xlsx"
REVIEW_DIR = OUTPUT_DIR / "review"


HEADERS = [
    "filename",
    "invoice_no",
    "vendor_name",
    "vendor_id",
    "invoice_date",
    "due_date",
    "currency",
    "netto",
    "ust",
    "brutto",
    "ust_rate",
    "category",
    "n_line_items",
    "processed_at",
]


def _ensure_workbook() -> Workbook:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if PROCESSED_XLSX.exists():
        return load_workbook(PROCESSED_XLSX)
    wb = Workbook()
    ws = wb.active
    ws.title = "processed"
    ws.append(HEADERS)
    wb.save(PROCESSED_XLSX)
    return load_workbook(PROCESSED_XLSX)


def append_processed_row(invoice: Invoice) -> None:
    wb = _ensure_workbook()
    ws = wb["processed"]
    ws.append(
        [
            invoice.filename,
            invoice.invoice_no,
            invoice.vendor_name,
            invoice.vendor_id or "",
            invoice.invoice_date.isoformat(),
            invoice.due_date.isoformat() if invoice.due_date else "",
            invoice.currency,
            invoice.netto,
            invoice.ust,
            invoice.brutto,
            invoice.ust_rate,
            invoice.category or "",
            len(invoice.line_items),
            datetime.now().isoformat(timespec="seconds"),
        ]
    )
    wb.save(PROCESSED_XLSX)


def write_review_entry(invoice: Invoice, reason: ReviewReason) -> Path:
    REVIEW_DIR.mkdir(parents=True, exist_ok=True)
    stem = Path(invoice.filename).stem
    out = REVIEW_DIR / f"{stem}__review.json"
    payload: dict[str, Any] = {
        "invoice": json.loads(invoice.model_dump_json()),
        "reason": json.loads(reason.model_dump_json()),
        "flagged_at": datetime.now().isoformat(timespec="seconds"),
    }
    out.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    return out


def existing_invoice_numbers() -> set[tuple[str, str]]:
    """Return the set of (invoice_no, vendor_id_or_name) tuples already in processed.xlsx."""
    if not PROCESSED_XLSX.exists():
        return set()
    wb = load_workbook(PROCESSED_XLSX, read_only=True)
    ws = wb["processed"]
    rows = ws.iter_rows(min_row=2, values_only=True)
    seen: set[tuple[str, str]] = set()
    for row in rows:
        # row layout matches HEADERS; invoice_no=idx 1, vendor_id=idx 3, vendor_name=idx 2
        if row and row[1]:
            key_id = str(row[3] or row[2] or "")
            seen.add((str(row[1]), key_id))
    return seen


def reset_outputs() -> None:
    """Wipe processed.xlsx and the review folder. For clean demo re-runs."""
    if PROCESSED_XLSX.exists():
        PROCESSED_XLSX.unlink()
    if REVIEW_DIR.exists():
        for f in REVIEW_DIR.iterdir():
            f.unlink()
